"""Export ``.xlsx`` da lista de checklists — ticket ``v1-entregavel/06``.

Custo de API: **zero**. Tudo aqui é leitura de banco e geração de planilha;
nenhuma chamada a OpenAI ou Anthropic acontece nestes testes.

Não reescreve a consulta: os mesmos helpers de fixture de
``test_portal_checklists.py`` (copiados aqui, não importados, para este
arquivo ficar independente de mudanças concorrentes em
``checklist_filter``/``checklist_query`` — ver colisão anotada no ticket).
"""

from __future__ import annotations

import uuid
from datetime import UTC, datetime
from io import BytesIO
from typing import Any

import bcrypt
import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session, sessionmaker
from sqlalchemy.pool import StaticPool

from app.core.config import AppEnv, Settings
from app.db.base import Base
from app.db.session import get_db
from app.main import create_app
from app.models.checklist_analysis import ChecklistViewResult
from app.models.pipeline import PipelineJob
from app.models.sisloc import SislocChecklist
from app.models.user import User

LISTA = "/api/v1/portal/checklists"
EXPORT = "/api/v1/portal/checklists/export.xlsx"
MEDIA_TYPE_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# ── fixtures ──────────────────────────────────────────────────────────────────


@pytest.fixture
def portal_settings() -> Settings:
    return Settings(
        _env_file=None,
        app_env=AppEnv.TEST,
        session_secret="test-secret-key-32-chars-minimum!",
    )


@pytest.fixture
def sqlite_engine():
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    yield engine
    Base.metadata.drop_all(engine)
    engine.dispose()


@pytest.fixture
def db(sqlite_engine) -> Session:
    factory = sessionmaker(bind=sqlite_engine, autocommit=False, autoflush=False)
    session = factory()
    try:
        yield session
    finally:
        session.close()


@pytest.fixture
def portal_client(portal_settings: Settings, db: Session) -> TestClient:
    def _override_db():
        yield db

    app = create_app(portal_settings)
    from app.core.config import get_settings

    app.dependency_overrides[get_settings] = lambda: portal_settings
    app.dependency_overrides[get_db] = _override_db
    return TestClient(app, raise_server_exceptions=False)


@pytest.fixture
def logado(portal_client: TestClient, db: Session) -> TestClient:
    hashed = bcrypt.hashpw(b"s3cr3t", bcrypt.gensalt()).decode()
    db.add(User(email="test@tecnogera.com", password_hash=hashed, is_active=True))
    db.commit()
    portal_client.post(
        "/api/v1/portal/login",
        json={"email": "test@tecnogera.com", "password": "s3cr3t"},
    )
    return portal_client


# ── helpers ───────────────────────────────────────────────────────────────────


def _snapshot(
    checklist_id: str,
    *,
    formulario: str = "F038 - PRÉ LOCAÇÃO DE GERADOR",
    filial: str | None = "MG-CGE",
    patrimonio: str | None = "TECG01364",
    projeto: str | None = "035514/2026-EBAZAR.COM.BR. LTDA",
    responsavel: str | None = "MATHEUS.PARAISO",
    data_conclusao: datetime | None = None,
    n_linhas: int = 1,
) -> dict[str, Any]:
    linha = SislocChecklist(
        codigo_checklist=checklist_id,
        formulario=formulario,
        filial=filial,
        patrimonio=patrimonio,
        projeto=projeto,
        responsavel=responsavel,
        data_conclusao=data_conclusao or datetime(2026, 8, 2, 14, 30, tzinfo=UTC),
        status="Concluído",
        origem="OM",
        numero_om=36729,
        ordem=1,
        n_linhas=n_linhas,
    )
    return linha.snapshot(lido_em=datetime(2026, 8, 2, 15, 0, tzinfo=UTC)).como_json()


def _job(
    db: Session,
    checklist_id: str,
    *,
    conformidade: str | None = "conforme",
    severidade: int | None = None,
    vista_determinante: str | None = None,
    vistas: str | None = "c54,c55,c56",
    status: str = "done",
    formulario: str = "F038 - PRÉ LOCAÇÃO DE GERADOR",
    filial: str | None = "MG-CGE",
    patrimonio: str | None = "TECG01364",
    projeto: str | None = "035514/2026-EBAZAR.COM.BR. LTDA",
    data_conclusao: datetime | None = None,
    n_linhas: int = 1,
    created_at: datetime | None = None,
) -> PipelineJob:
    job = PipelineJob(
        id=uuid.uuid4(),
        checklist_id=checklist_id,
        status=status,
        mode="sync",
        conformidade=conformidade,
        severidade_max=severidade,
        vista_determinante=vista_determinante,
        vistas_recebidas=vistas,
        formulario=formulario,
        patrimonio=patrimonio,
        projeto=projeto,
        n_linhas=n_linhas,
        llm_cost_usd=0.002,
        llm_calls=3,
        created_at=created_at or datetime(2026, 8, 2, 16, 0, tzinfo=UTC),
        sisloc_snapshot=_snapshot(
            checklist_id,
            formulario=formulario,
            filial=filial,
            patrimonio=patrimonio,
            projeto=projeto,
            data_conclusao=data_conclusao,
            n_linhas=n_linhas,
        ),
    )
    db.add(job)
    db.commit()
    return job


def _nao_conforme(db: Session, checklist_id: str, **kwargs: Any) -> PipelineJob:
    job = _job(
        db,
        checklist_id,
        conformidade="nao_conforme",
        severidade=kwargs.pop("severidade", 2),
        vista_determinante="c54",
        **kwargs,
    )
    db.add(
        ChecklistViewResult(
            id=uuid.uuid4(),
            job_id=job.id,
            checklist_id=job.checklist_id,
            campo="c54",
            dropbox_path=f"/Sisloc/MG-CGE/{job.checklist_id} 01/c54 foto.jpg",
            status="analisada",
            conformidade="nao_conforme",
            vista_confere=True,
            achados=[
                {
                    "classe": "dano_visivel",
                    "tipo_defeito": "amassado_deformacao",
                    "severidade": 2,
                    "local": "quadrante inferior direito",
                    "observacao": "Amassado visível.",
                    "confianca": 0.87,
                }
            ],
            severidade_max=2,
            classe="dano_visivel",
            tipo_defeito="amassado_deformacao",
            confianca=0.87,
            model_version="gpt-4.1-mini",
            cost_usd=0.002,
        )
    )
    db.commit()
    return job


def _abrir_planilha(conteudo: bytes):
    return load_workbook(BytesIO(conteudo))


_CABECALHO_ESPERADO = (
    "ID checklist",
    "Ativo (patrimônio)",
    "Cliente",
    "Filial",
    "Formulário",
    "Indicador",
    "Severidade",
    "Vista determinante",
    "Validação",
    "Data de conclusão",
    "Processado em",
)


# ── autenticação ──────────────────────────────────────────────────────────────


def test_export_requer_autenticacao(portal_client):
    assert portal_client.get(EXPORT).status_code == 401


# ── contrato HTTP ─────────────────────────────────────────────────────────────


def test_export_content_type_e_content_disposition(logado, db):
    _job(db, "311989")

    resp = logado.get(EXPORT)

    assert resp.status_code == 200
    assert resp.headers["content-type"] == MEDIA_TYPE_XLSX
    disposition = resp.headers["content-disposition"]
    assert disposition.startswith("attachment;")
    assert ".xlsx" in disposition
    # nome do arquivo inclui a data da exportação — pedido do ticket.
    hoje = datetime.now(UTC).strftime("%Y-%m-%d")
    assert hoje in disposition


def test_export_abre_como_xlsx_valido_com_uma_aba(logado, db):
    _job(db, "311989")

    resp = logado.get(EXPORT)
    wb = _abrir_planilha(resp.content)

    # Uma aba só — a aba de achados (uma linha por vista) foi recusada no
    # definição de produto. A planilha é espelho da tela, não tabela dinâmica.
    assert wb.sheetnames == ["Checklists"]


def test_export_cabecalho_em_negrito_com_auto_filter(logado, db):
    _job(db, "311989")

    resp = logado.get(EXPORT)
    ws = _abrir_planilha(resp.content)["Checklists"]

    valores_cabecalho = tuple(c.value for c in ws[1])
    assert valores_cabecalho == _CABECALHO_ESPERADO
    assert all(c.font.bold for c in ws[1])
    assert ws.auto_filter.ref is not None


def test_export_422_indicador_invalido(logado, db):
    assert logado.get(EXPORT, params={"indicador": "quase_conforme"}).status_code == 422


def test_export_422_validacao_invalida(logado, db):
    assert logado.get(EXPORT, params={"validacao": "meio_certo"}).status_code == 422


def test_export_422_ordenar_invalido(logado, db):
    assert logado.get(EXPORT, params={"ordenar": "custo"}).status_code == 422


# ── grão e recorte ───────────────────────────────────────────────────────────


def test_export_uma_linha_por_checklist(logado, db):
    _nao_conforme(db, "1")
    _job(db, "2", conformidade="conforme")
    _job(db, "3", conformidade="nao_processavel")

    resp = logado.get(EXPORT)
    ws = _abrir_planilha(resp.content)["Checklists"]

    assert ws.max_row == 1 + 3  # cabeçalho + 3 checklists


def test_export_ignora_limit_e_exporta_todo_o_conjunto_filtrado(logado, db):
    """Mais linhas que o `limit` default (50) da lista — critério de aceite do ticket."""
    for i in range(60):
        _job(db, f"c{i:03d}", conformidade="conforme")

    # a lista, sem `limit` explícito, devolve só a primeira página
    lista = logado.get(LISTA).json()
    assert lista["total"] == 60
    assert len(lista["itens"]) == 50  # default da paginação

    resp = logado.get(EXPORT)
    ws = _abrir_planilha(resp.content)["Checklists"]

    assert ws.max_row == 1 + 60  # export não pagina


def test_export_filtro_de_filial_chega_na_planilha(logado, db):
    _job(db, "1", filial="MG-CGE")
    _job(db, "2", filial="MG-CGE")
    _job(db, "3", filial="SP-GRU")

    resp = logado.get(EXPORT, params={"filial": "MG-CGE"})
    ws = _abrir_planilha(resp.content)["Checklists"]

    assert ws.max_row == 1 + 2
    filiais_na_planilha = {ws.cell(row=r, column=4).value for r in range(2, ws.max_row + 1)}
    assert filiais_na_planilha == {"MG-CGE"}


def test_export_filtro_de_indicador_chega_na_planilha(logado, db):
    _nao_conforme(db, "1")
    _job(db, "2", conformidade="conforme")

    resp = logado.get(EXPORT, params={"indicador": "nao_conforme"})
    ws = _abrir_planilha(resp.content)["Checklists"]

    assert ws.max_row == 1 + 1
    assert ws.cell(row=2, column=1).value == "1"


# ── rótulos e tipos de célula ────────────────────────────────────────────────


def test_export_nenhum_valor_cru_de_enum(logado, db):
    _nao_conforme(db, "1")

    resp = logado.get(EXPORT)
    ws = _abrir_planilha(resp.content)["Checklists"]

    linha = {ws.cell(row=1, column=c).value: ws.cell(row=2, column=c).value for c in range(1, 12)}
    assert linha["Indicador"] == "Não conforme"
    assert linha["Indicador"] != "nao_conforme"
    assert linha["Severidade"] == "Alta"
    assert linha["Vista determinante"] == "Lateral direita"
    # `validacao` é sempre `pendente` até o ticket 10 — o rótulo de tela é
    # "A validar", nunca o valor cru do enum de processo.
    assert linha["Validação"] == "A validar"
    assert linha["Validação"] != "pendente"


def test_export_datas_sao_celulas_de_data_nao_string(logado, db):
    _job(db, "1", data_conclusao=datetime(2026, 8, 2, 14, 30, tzinfo=UTC))

    resp = logado.get(EXPORT)
    ws = _abrir_planilha(resp.content)["Checklists"]

    celula_data = ws.cell(row=2, column=10)
    celula_processado = ws.cell(row=2, column=11)
    assert isinstance(celula_data.value, datetime)
    assert isinstance(celula_processado.value, datetime)
    assert celula_data.number_format != "General"


def test_export_severidade_nula_vira_traco_ou_vazio_nunca_numero_cru(logado, db):
    _job(db, "1", conformidade="conforme")

    resp = logado.get(EXPORT)
    ws = _abrir_planilha(resp.content)["Checklists"]

    assert ws.cell(row=2, column=7).value is None


# ── consistência com a tela ──────────────────────────────────────────────────


def test_export_respeita_a_mesma_ordenacao_padrao_da_lista(logado, db):
    _job(db, "baixa", conformidade="nao_conforme", severidade=4)
    _job(db, "critica", conformidade="nao_conforme", severidade=1)
    _job(db, "alta", conformidade="nao_conforme", severidade=2)

    lista = logado.get(LISTA).json()
    ids_lista = [i["checklist_id"] for i in lista["itens"]]

    resp = logado.get(EXPORT)
    ws = _abrir_planilha(resp.content)["Checklists"]
    ids_planilha = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]

    assert ids_planilha == ids_lista == ["critica", "alta", "baixa"]


def test_export_vazio_gera_planilha_so_com_cabecalho(logado, db):
    resp = logado.get(EXPORT, params={"filial": "NENHUMA-FILIAL"})

    assert resp.status_code == 200
    ws = _abrir_planilha(resp.content)["Checklists"]
    assert ws.max_row == 1
