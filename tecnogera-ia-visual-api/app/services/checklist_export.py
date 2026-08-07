"""Export ``.xlsx`` da lista de checklists — ticket ``v1-entregavel/06``.

Espelho da tela: **uma linha por checklist**, **todo o conjunto filtrado**
(ignora ``limit``/``offset``), nenhum valor cru de enum — sempre o
``*_rotulo`` em português. Aba de achados (uma linha por vista) foi recusada
na definição de produto; se a Tecnogera pedir uma tabela dinâmica por achado, isso é
ticket novo.

Reusa ``checklist_query.listar_checklists`` em vez de montar consulta
própria: a planilha não pode discordar da tela. Este módulo não sabe nada de
SQL — só formata o que a consulta já devolve.

``.xlsx`` de verdade via ``openpyxl``, não CSV: o Excel pt-BR interpreta
``,`` como separador decimal e espera ``;`` como separador de campo em CSV, e
qualquer acento sem BOM correto quebra na abertura. ``.xlsx`` binário não tem
essa armadilha.
"""

from __future__ import annotations

from dataclasses import replace
from datetime import UTC, datetime
from io import BytesIO
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.services import checklist_query as cq

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet
    from sqlalchemy.orm import Session

#: "Não invente teto" (ticket) — ~71 checklists/mês depois do corte para
#: F038, então mesmo anos de histórico cabem longe deste número. Um limite
#: alto e único evita duas idas ao banco só para descobrir o total antes de
#: buscar tudo.
_LIMITE_EXPORT = 1_000_000

#: `validacao` não tem `*_rotulo` no contrato — é enum de PROCESSO, não do
#: laudo (ver `checklist_query` e o comentário equivalente no front,
#: `ChecklistsPage.tsx`). O rótulo de apresentação mora aqui, espelhando
#: exatamente o que a tela já mostra na coluna "Validação", para a planilha
#: não inventar um vocabulário paralelo.
ROTULO_VALIDACAO: dict[str, str] = {
    "pendente": "A validar",
    "confirmado": "Confirmado",
    "corrigido": "Corrigido",
}

#: (título da coluna, largura aproximada em caracteres). Espelha a tela
#: (ID checklist, Ativo, Filial, Indicador, Sev., Vista, Data, Validação)
#: mais o que só faz sentido numa planilha para consumo externo: Cliente,
#: Formulário por extenso e o timestamp de processamento da esteira.
_COLUNAS: tuple[tuple[str, int], ...] = (
    ("ID checklist", 14),
    ("Ativo (patrimônio)", 20),
    ("Cliente", 34),
    ("Filial", 10),
    ("Formulário", 28),
    ("Indicador", 18),
    ("Severidade", 12),
    ("Vista determinante", 20),
    ("Validação", 14),
    ("Data de conclusão", 20),
    ("Processado em", 20),
)

_FORMATO_DATA = "DD/MM/YYYY HH:MM"


def gerar_planilha(db: Session, filtros: cq.ChecklistFiltros) -> BytesIO:
    """Gera o ``.xlsx`` do conjunto FILTRADO — ignora ``limit``/``offset``.

    Qualquer ``limit``/``offset`` recebido em ``filtros`` é substituído: o
    export é "tudo que a tela mostraria sem paginação", nunca uma página. Os
    demais filtros (indicador, validação, filial, formulário,
    ``codigo_checklist``, período, ordenação) são os mesmos que a tela usa —
    a chamada é a mesma consulta, só sem limite de página.
    """
    filtros_export = replace(filtros, limit=_LIMITE_EXPORT, offset=0)
    pagina = cq.listar_checklists(db, filtros_export)

    livro = Workbook()
    planilha = livro.active
    assert planilha is not None  # `Workbook()` sempre cria a aba ativa
    planilha.title = "Checklists"

    _escrever_cabecalho(planilha)
    linha_atual = 2
    for item in pagina.itens:
        _escrever_linha(planilha, linha_atual, item)
        linha_atual += 1

    # `auto_filter` na primeira linha e datas como célula de data (não texto)
    # são o que permite ordenar/filtrar no Excel — pedido explícito do
    # ticket, não cosmética.
    planilha.auto_filter.ref = planilha.dimensions
    planilha.freeze_panes = "A2"

    buffer = BytesIO()
    livro.save(buffer)
    buffer.seek(0)
    return buffer


def nome_arquivo(agora: datetime | None = None) -> str:
    """Nome do anexo, com a data da exportação — pedido explícito do ticket."""
    momento = agora or datetime.now(UTC)
    return f"checklists-{momento:%Y-%m-%d}.xlsx"


def _escrever_cabecalho(planilha: Worksheet) -> None:
    for indice, (titulo, largura) in enumerate(_COLUNAS, start=1):
        celula = planilha.cell(row=1, column=indice, value=titulo)
        celula.font = Font(bold=True)
        planilha.column_dimensions[get_column_letter(indice)].width = largura


def _sem_fuso(valor: datetime | None) -> datetime | None:
    """``openpyxl``/Excel não aceitam ``tzinfo`` em célula de data.

    Em produção (Postgres) as datas do job vêm com timezone; o contrato já
    documenta "trate como UTC" para quem consome — aqui é o mesmo trato,
    convertendo para UTC antes de descartar o fuso, para não silenciosamente
    mudar a hora de quem gravou em outro fuso.
    """
    if valor is None:
        return None
    if valor.tzinfo is not None:
        valor = valor.astimezone(UTC)
    return valor.replace(tzinfo=None)


def _escrever_linha(planilha: Worksheet, linha: int, item: cq.ChecklistLinha) -> None:
    valores: tuple[str | datetime | None, ...] = (
        item.checklist_id,
        item.patrimonio,
        item.cliente,
        item.filial,
        item.formulario,
        item.indicador_rotulo,
        item.severidade_rotulo,
        item.vista_determinante_rotulo,
        ROTULO_VALIDACAO.get(item.validacao, item.validacao),
        _sem_fuso(item.data),
        _sem_fuso(item.criado_em),
    )
    for coluna, valor in enumerate(valores, start=1):
        celula = planilha.cell(row=linha, column=coluna, value=valor)
        if isinstance(valor, datetime):
            celula.number_format = _FORMATO_DATA


__all__ = ["ROTULO_VALIDACAO", "gerar_planilha", "nome_arquivo"]
